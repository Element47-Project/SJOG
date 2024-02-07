import os
import io
import pandas as pd
import pyodbc
from dotenv import load_dotenv
import time
from exchangelib import Credentials, Account, DELEGATE, FileAttachment
import openpyxl
import xlrd
from Gas_csv_Formatting import consumption
from fuzzywuzzy import process
import pdfplumber
import requests
from datetime import datetime, timedelta

# Load environment variables
load_dotenv()
''' The PDF file path configuration. Written by Prachi'''
# Get the path for this script to save pdfs
script_path = __file__
# To get the absolute path to the script file, use abspath
absolute_script_path = os.path.abspath(__file__)
# To get the directory containing the script, use dir name
script_dir = os.path.dirname(absolute_script_path)
pdf_dir = script_dir + '/pdfs'
PDF_DIR = '/pdfs'

# All settings in the .env file, including SQL and Email information.
EMAIL_ADDRESS = os.environ.get("EMAIL_ADDRESS")
DESIRED_DOMAINS = ['@gmail.com']
PASSWORD = os.environ.get('PASSWORD')
SQL_SERVER = os.environ.get('AZURE_SQL_SERVER')
SQL_DB_NAME = os.environ.get('AZURE_SQL_DB_NAME')
SQL_USERNAME = os.environ.get('AZURE_SQL_USERNAME')
SQL_PASSWORD = os.environ.get('AZURE_SQL_PASSWORD')
CONNECTION_STRING = (
    f'DRIVER={{ODBC Driver 18 for SQL Server}};SERVER={SQL_SERVER};'
    f'DATABASE={SQL_DB_NAME};UID={SQL_USERNAME};PWD={SQL_PASSWORD}'
)


def connect_to_db(conn_str):
    """Establishes a connection to the database."""
    with pyodbc.connect(conn_str) as conn:
        cursor = conn.cursor()
    return conn, cursor


def get_all_table_columns(cursor):
    tables_columns = {}
    cursor.execute("SELECT TABLE_NAME, COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS ORDER BY TABLE_NAME, "
                   "ORDINAL_POSITION")
    for row in cursor.fetchall():
        table_name, column_name = row
        if table_name not in tables_columns:
            tables_columns[table_name] = []
        tables_columns[table_name].append(column_name)
    return tables_columns


def get_all_table_primary_keys(cursor):
    """
    Retrieves all primary key columns for each table in the database.
    """
    # Retrieve all base tables
    cursor.execute("SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE = 'BASE TABLE'")
    table_names = [row.TABLE_NAME for row in cursor.fetchall()]
    # Dictionary to hold table primary key information
    primary_keys = {}

    for table_name in table_names:
        # Query to find primary key columns for the current table
        pk_query = """
            SELECT COLUMN_NAME
            FROM INFORMATION_SCHEMA.KEY_COLUMN_USAGE
            WHERE TABLE_NAME = ? AND OBJECTPROPERTY(OBJECT_ID(CONSTRAINT_SCHEMA + '.' + CONSTRAINT_NAME), 'IsPrimaryKey') = 1
            ORDER BY ORDINAL_POSITION
        """
        cursor.execute(pk_query, [table_name])
        pk_columns = cursor.fetchall()
        # Add table primary key info to the dictionary
        primary_keys[table_name] = [col.COLUMN_NAME for col in pk_columns]
    return primary_keys, table_names


def find_header_row(df, expected_columns):
    """
    Finds the header row index in a DataFrame by matching it to expected column names.
    """
    expected_columns_set = set(expected_columns)

    for i, row in df.iterrows():
        # Extract non-null values from the row
        row_values_set = set(row.dropna())
        if row_values_set == expected_columns_set:
            print(f"Matching header row found at index: {i}")
            return i
    return None


def process_xlsx_attachments(attachment, table_dict, cursor, conn):
    all_tables_columns = get_all_table_columns(cursor)
    try:
        # Open the workbook in memory
        excel_stream = io.BytesIO(attachment.content)
        workbook = openpyxl.load_workbook(excel_stream, read_only=True)

        excel_stream.seek(0)  # Ensure we are at the beginning of the stream
        first_sheet = workbook[workbook.sheetnames[0]]  # Use the first sheet to identify header and table
        temp_df = pd.DataFrame(first_sheet.iter_rows(values_only=True, max_row=20))
        header_row_index = None
        upload_table = None
        for table_name, azure_columns in all_tables_columns.items():
            header_row_index = find_header_row(temp_df, azure_columns)
            if header_row_index is not None:
                upload_table = table_name
                break  # Found a matching header, break out of the loop

        # If header_row_index or table name not found, return or raise an exception
        if header_row_index is None or upload_table is None:
            print(f"Header row or table name not found in the Excel file {attachment.name}")
            return

        # Process each sheet using the found header row index and table name
        for sheet_name in workbook.sheetnames:
            excel_stream.seek(0)  # Reset the stream to the beginning for each sheet

            # Read the entire sheet starting from the header row index
            batch_df = pd.read_excel(excel_stream, sheet_name=sheet_name, skiprows=header_row_index)
            batch_df.columns = temp_df.iloc[header_row_index].values  # Set the correct header

            # Upload the DataFrame to the database
            upload_dataframe_to_azure_sql(batch_df, upload_table, cursor, table_dict, conn)

    except xlrd.biffh.XLRDError as e:
        if str(e) == "Workbook is encrypted":
            print(f"Cannot process encrypted file: {attachment.name}")
        else:
            raise


def process_csv_attachments(attachment, table_dict, cursor, conn):
    file_name_without_extension = attachment.name.rsplit('.', 1)[0]
    csv_header = pd.read_csv(io.BytesIO(attachment.content), nrows=0).columns.tolist()
    if file_name_without_extension in table_dict:
        csv_data = pd.read_csv(io.BytesIO(attachment.content))
        upload_dataframe_to_azure_sql(csv_data, file_name_without_extension, cursor, table_dict, conn)
    elif csv_header == ['NMI', 'CHECKSUM', 'GASDAY', 'READ_TYPE', 'DAILY_HEAT_VALUE',
                        'CONSUMPTION_HR01', 'CONSUMPTION_HR02', 'CONSUMPTION_HR03', 'CONSUMPTION_HR04',
                        'CONSUMPTION_HR05', 'CONSUMPTION_HR06', 'CONSUMPTION_HR07', 'CONSUMPTION_HR08',
                        'CONSUMPTION_HR09', 'CONSUMPTION_HR10', 'CONSUMPTION_HR11', 'CONSUMPTION_HR12',
                        'CONSUMPTION_HR13', 'CONSUMPTION_HR14', 'CONSUMPTION_HR15', 'CONSUMPTION_HR16',
                        'CONSUMPTION_HR17', 'CONSUMPTION_HR18', 'CONSUMPTION_HR19', 'CONSUMPTION_HR20',
                        'CONSUMPTION_HR21', 'CONSUMPTION_HR22', 'CONSUMPTION_HR23', 'CONSUMPTION_HR24',
                        'TOTAL_DAILY_CONSUMPTION', 'PEAK_RATE']:
        csv_data = pd.read_csv(io.BytesIO(attachment.content))
        df_csv = consumption(csv_data)
        upload_dataframe_to_azure_sql(df_csv, 'TestingGas', cursor, table_dict, conn)


def process_pdf_tables(attachment_content, directory=pdf_dir, filename=None):
    # Ensure the directory exists
    if not os.path.isdir(directory):
        os.makedirs(directory, exist_ok=True)
    # If no filename is given, generate one with a timestamp
    if filename is None:
        filename = f"pdf_table_{int(time.time())}.pdf"
    file_path = os.path.join(directory, filename)
    # Save the PDF to the specified directory
    with open(file_path, "wb") as f:
        f.write(attachment_content.read())

    dataframes = []
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            # Extract table from the page
            table = page.extract_table()
            if table:
                # Convert table (list of lists) to DataFrame
                df = pd.DataFrame(table[1:], columns=table[0])
                df.columns = [col.replace('\n', ' ').strip().title() for col in df.columns]  # Clean column names
                dataframes.append(df)

    print(f"Number of tables extracted: {len(dataframes)}")

    for i, df in enumerate(dataframes):
        print(f"Preview of table {i}:")
        print(df.head())  # Print first few rows of each table

    # Optionally, you can delete the PDF file after processing
    # os.remove(file_path)
    return dataframes


def process_pdf_attachments(attachment, table_dict, cursor, conn):
    filename = attachment.name.rsplit('.', 1)[0]
    all_tables_columns = get_all_table_columns(conn)
    try:
        pdf_dataframes = process_pdf_tables(io.BytesIO(attachment.content), filename=f"{filename}.pdf")
        for pdf_df in pdf_dataframes:
            # Clean column names by replacing newlines and extra spaces
            pdf_df.columns = [col.replace('\n', ' ').strip().title() for col in pdf_df.columns]
            # print(f"Expected columns for {table_name}: {azure_columns}")
            table_name = ""
            # Implement fuzzy matching here
            for col in list(pdf_df.columns):  # Use list to avoid iterating over changing dict
                for table_name, azure_columns in all_tables_columns.items():
                    best_match = find_best_match(col, azure_columns)
                    if best_match:
                        pdf_df.rename(columns={col: best_match}, inplace=True)
                        return table_name

            # Upload the processed data to Azure SQL
            # Note: Ensure that the pdf_df now aligns with the table structure of Azure SQL
            upload_dataframe_to_azure_sql(pdf_df, table_name, cursor, table_dict, conn)
            # print("Data uploaded successfully to Azure SQL.")
    except Exception as e:
        print(f"Error processing PDF tables in file: {attachment.name}. Error: {e}")


def find_best_match(header, choices, threshold=95):
    best_match, score = process.extractOne(header, choices)
    return best_match if score >= threshold else None


def process_email_attachments(cursor, attachment_files, table_dict, conn):
    for item in attachment_files:
        if item.attachments:
            for attachment in item.attachments:
                print(f"Processing attachment: {attachment.name}")
                filename, extension = os.path.splitext(attachment.name)
                if isinstance(attachment, FileAttachment):
                    if extension in ['.xlsx', '.xls']:
                        process_xlsx_attachments(attachment, table_dict, cursor, conn)
                    elif extension == '.csv':
                        process_csv_attachments(attachment, table_dict, cursor, conn)
                    elif extension == '.pdf':
                        process_pdf_attachments(attachment, table_dict, cursor, conn)

                # Mark the item as read after processing
                item.is_read = True


def upload_dataframe_to_azure_sql(df, table_name, cursor, table_dict, conn):
    print(f"Uploading data to {table_name}. Please wait...")
    primary_keys = table_dict[table_name]

    if len(primary_keys) == 1:
        # Single primary key scenario
        pk_col = primary_keys[0]
        cursor.execute(f"SELECT TOP 1 [{pk_col}] FROM [{table_name}] ORDER BY [{pk_col}] DESC")
        last_record = cursor.fetchone()
        last_value = last_record[0] if last_record else None
        df[pk_col] = pd.to_datetime(df[pk_col], format='%d-%b-%Y %H:%M:%S', errors='coerce')
        if last_value is not None:
            df = df[df[pk_col] > last_value].copy()
        df[pk_col] = df[pk_col].dt.strftime('%Y-%m-%d %H:%M:%S').astype(str)

    elif len(primary_keys) == 2:
        cursor.execute(f"""
                SELECT [NMI], MAX([END INTERVAL]) 
                FROM [{table_name}] 
                GROUP BY [NMI]
                """)
        last_records = {nmi: max_end_interval for nmi, max_end_interval in cursor.fetchall()}
        filtered_df = pd.DataFrame()
        for nmi, last_time in last_records.items():
            temp_df = df[(df['NMI'].astype(str) == str(nmi)) & (df['END INTERVAL'] > last_time)]
            filtered_df = pd.concat([filtered_df, temp_df], ignore_index=True)
        df = filtered_df
    if df.empty:
        print("No new rows to insert after filtering with last records.")
        return

    # Perform batch insert
    data_for_insert = [tuple(None if pd.isnull(item) else item for item in row) for row in df.to_records(index=False)]
    batch_insert(cursor, table_name, df.columns.tolist(), data_for_insert, conn)


def batch_insert(cursor, table_name, columns, data, conn, batch_size=1000):
    """
    Inserts data into a database in batches.
    """
    sql_columns = ', '.join([f'[{col}]' for col in columns])
    placeholders = ', '.join(['?'] * len(columns))
    insert_query = f"INSERT INTO [{table_name}] ({sql_columns}) VALUES ({placeholders})"

    for start in range(0, len(data), batch_size):
        batch = data[start:start + batch_size]
        try:
            cursor.executemany(insert_query, batch)
            conn.commit()
            print(f'Batch {start//batch_size + 1}/{(len(data) + batch_size - 1) // batch_size} processed.')
        except pyodbc.Error as e:
            print(f"Error occurred in batch {start//batch_size + 1}: {e}")
            conn.rollback()
            if batch_size > 100:
                smaller_batch_size = max(batch_size // 2, 100)
                batch_insert(cursor, table_name, columns, batch, conn, smaller_batch_size)


def fetch_latest_date_from_azure(cursor, table_name):
    query = f"SELECT MAX(Date_time) FROM {table_name}"
    cursor.execute(query)
    result = cursor.fetchone()
    if result[0] is not None:
        return result[0]
    else:
        return None


def fetch_weather_data(latitude, longitude, start_date, end_date):
    url = "https://archive-api.open-meteo.com/v1/archive"
    params = {
        "latitude": latitude,
        "longitude": longitude,
        "start_date": start_date,
        "end_date": end_date,
        "hourly": "temperature_2m",
        "timezone": "auto"
    }
    response = requests.get(url, params=params)
    if response.status_code == 200:
        return response.json()
    else:
        return None


def process_weather_data(weather_data):
    # Accessing nested data within 'hourly' key
    hourly_data = weather_data['hourly']
    hourly_times = hourly_data['time']
    hourly_temperatures = hourly_data['temperature_2m']

    # Convert the ISO8601 time strings to datetime objects
    times = pd.to_datetime(hourly_times)

    # Create the DataFrame
    df_weather = pd.DataFrame({'Date_Time': times, 'Temperature': hourly_temperatures})

    return df_weather


def upload_temperature_to_azure_sql(df, table_name, conn, cursor, batch_size=2000):
    print(f"Uploading data to {table_name}")
    df_columns = df.columns.tolist()
    sql_columns = ', '.join([f'[{col}]' for col in df_columns])
    placeholders = ', '.join(['?'] * len(df_columns))
    insert_query = f"INSERT INTO {table_name} ({sql_columns}) VALUES ({placeholders})"
    data_for_insert = [tuple(row) for row in df.itertuples(index=False, name=None)]
    try:
        # Perform batch insert
        batch_insert(cursor, table_name, df_columns, data_for_insert, conn)
        conn.commit()
    except pyodbc.Error as e:
        print(f"Error during batch insert: {e}")
        conn.rollback()


def main():
    # Set up the email account
    print("Connecting to SQL Database...")
    credentials = Credentials(EMAIL_ADDRESS, PASSWORD)
    account = Account(
        EMAIL_ADDRESS,
        credentials=credentials,
        autodiscover=True,
        access_type=DELEGATE
    )
    conn, cursor = connect_to_db(CONNECTION_STRING)
    print("Connected. Loading the information from Database...")
    table_dict, all_tables = get_all_table_primary_keys(cursor)
    # Process unread emails
    all_unread_emails = account.inbox.filter(is_read=False).order_by('-datetime_received')
    filtered_unread_emails = [
        email for email in all_unread_emails
        if email.sender and email.sender.email_address and
        any(email.sender.email_address.strip().lower().endswith(domain) for domain in DESIRED_DOMAINS)
    ]
    process_email_attachments(cursor, filtered_unread_emails, table_dict, conn)

    # Upload the Temperature data
    tem_input = input("Do you want to upload the temperature data to Azure SQL? (Y to continue / N to stop): ")
    if tem_input == 'Y' or 'y' or '':
        table_name = 'Temperature_hourly'
        latest_date = fetch_latest_date_from_azure(cursor, table_name)
        if latest_date:
            start_date = (latest_date - timedelta(days=1 / 3)).strftime("%Y-%m-%d")
        else:
            start_date = "2020-01-01"  # Default start date if no latest date is available
        end_date = (datetime.now() - timedelta(days=3)).strftime("%Y-%m-%d")
        latitude = -31.9522
        longitude = 115.8614
        weather_data = fetch_weather_data(latitude, longitude, start_date, end_date)

        if weather_data:
            df_weather = process_weather_data(weather_data)
            upload_temperature_to_azure_sql(df_weather, table_name, conn, cursor)
            print("Weather data upload complete.")
        else:
            print("Failed to fetch weather data.")
    else:
        print("Process is closing.")

    # Upload the time into SQL table 'last_read_time'
    now = datetime.now()
    formatted_datetime = now.strftime('%Y-%m-%dT%H:%M:%S')
    inset_query_time = 'INSERT INTO Last_read_time (Read_time) VALUES (?)'
    cursor.execute(inset_query_time, formatted_datetime)
    conn.commit()
    conn.close()


if __name__ == "__main__":
    main()
